import sys
import time


from openpyxl import Workbook, load_workbook
print()
print()
print("""This program is designed to copy ranges of cells from various sheets 
in source Excel file, to one list in destination file for further usage. 
It can modify files with ".xlsx" extension only.

If the file is located in the same path (folder) as source file, enter name of
source file (with extension). Otherwise enter full path to the file (with extension)
""")


S_FILENAME = input("====> :")
S_FILEEXTENSION = "" #".xlsx"
SOURCE_XL = load_workbook(S_FILENAME+S_FILEEXTENSION)
print()
print("Source file has the following sheets: ")

# printing all file's sheets and options

SHEET_POSITION = 0
for _ in SOURCE_XL.sheetnames:

    print( SHEET_POSITION + 1, "..........", SOURCE_XL.sheetnames[SHEET_POSITION])
    SHEET_POSITION +=1

# printing all options

print("total number of sheets found", SHEET_POSITION)
print()
print("a .......... SELECT ALL")
print("s .......... SAVE SHEET'S LIST")
print("e .......... EXIT PROGRAM")
print()
print("""Please select ordinal number of a sheet, You wish to copy data from.
One at a time and press ENTER
or select option letter""")

# Handling user input

CHOOSEN_OPTION = input("Choose option: ")
CHOOSEN_SHEETS = set()

if CHOOSEN_OPTION == "e":
    print("Exiting program")
    time.sleep(1.0)
    sys.exit(0)

elif CHOOSEN_OPTION == "a":
    CHOOSEN_SHEETS = SOURCE_XL.sheetnames

elif  True:
    try:
        while (CHOOSEN_OPTION != "s"):
            CHOOSEN_SHEETS.add(SOURCE_XL.sheetnames[int(CHOOSEN_OPTION) - 1])
            print("Added ", SOURCE_XL.sheetnames[int(CHOOSEN_OPTION) - 1], "to set successfully")
            print("Sheets in set: ", CHOOSEN_SHEETS)
            CHOOSEN_OPTION = input("Choose option: ")
    except :
        print("Unknown command!")
        time.sleep(1.1)

print("You have successfully choosen ",CHOOSEN_SHEETS, "to copy" )
print("Total number of choosen sheets is ", len(CHOOSEN_SHEETS))

# Creating new workbook
# Creating new sheets - "summary" and "conv_sheet"

# choosing output file

print("""enter name of destination file (with file extension) 
and destination sheet title separated by a coma""")

D_FILENAME,D_SHEETNAME = input("====> :").split(",")
D_FILEEXTENSION = "" #".xlsx"
DEST_XL = load_workbook(D_FILENAME+D_FILEEXTENSION)



DEST_XL.create_sheet("summary")
SUM_SHEET = DEST_XL["summary"]
# SUM_SHEET.title = "summary"
# DEST_XL.create_sheet("conv_list")
SUM_SHEET['A1'].value = "List of converted sheets:"

A = 2
for _ in CHOOSEN_SHEETS:
    SUM_SHEET.cell(row=A, column=1).value = _
    A += 1

CHOOSEN_SHEETS = list(CHOOSEN_SHEETS)

# Defining function for copying range of cells from source excel sheet, to destination excel sheet
# appending sheet name to first column in destination excel sheet

def sheetcopy(SOURCE_SH, SROW_S, SROW_E, SCOL_S, SCOL_E, DROW, DCOL):
    COLUMNS_DONE = 0
    while SCOL_S <= SCOL_E:
        COLUMNS_DONE += 1
        ROWS_DONE = 0
        while SROW_S <= SROW_E:
            CONV_SHEET.cell(row=DROW, column=DCOL).value = SOURCE_SH.cell(row=SROW_S, column=SCOL_S).value
            DROW += 1
            SROW_S += 1
            ROWS_DONE += 1
        # print("rows done :", ROWS_DONE)
        DROW = DROW - ROWS_DONE
        SROW_S = SROW_S - ROWS_DONE
        SCOL_S += 1
        DCOL += 1
    while SROW_S <= SROW_E:
        CONV_SHEET.cell(row=DROW, column=DCOL - COLUMNS_DONE - 1).value = SOURCE_SH.title
        DROW += 1
        SROW_S += 1
        ROWS_DONE += 1
    SCOL_S = SCOL_S - COLUMNS_DONE
    DCOL = DCOL - COLUMNS_DONE




    return (SOURCE_SH, SROW_S, SROW_E, SCOL_S, SCOL_E, DROW, DCOL)

def sheetcheck(SOURCE_SH, ROW_CHECK, COL_CHECK):
    while SOURCE_SH.cell(row=ROW_CHECK, column=COL_CHECK).value != None :
        ROW_CHECK += 1
    return (SOURCE_SH, ROW_CHECK, COL_CHECK)

# Executing function and returning end row in destination excel sheet
print("""Do you want to edit source and destination copy ranges
if so type "y" if no type any other character and press enter. Default options are:
(Source: starting row no 5, starting column 3, ending column 14, Destination: row start no 2, column start no 2)
""")

RANGES = input("====> :")

if RANGES != "y":
    SROW_S = 5
    SROW_E = 23
    SCOL_S = 3
    SCOL_E = 14
    DROW = 2
    DCOL = 2
    CONV_SHEET = DEST_XL[D_SHEETNAME]
    SOURCE_SH = SOURCE_XL[CHOOSEN_SHEETS[0]]
    ROW_CHECK = SROW_S
    COL_CHECK = SCOL_S
else:
    print("""Type positional numbers in following sequence:
    (Source: starting row, starting column, ending column, Destination: row start no, column start no)
    or example 5,3,14,2,2 where 5 is row 5, and 3 is column C """)
    SROW_S,SCOL_S,SCOL_E,DROW,DCOL = input("====> :").split(",")
    SROW_S = int(SROW_S)
    SCOL_S = int(SCOL_S)
    SCOL_E = int(SCOL_E)
    DROW = int(DROW)
    DCOL = int(DCOL)
    CONV_SHEET = DEST_XL[D_SHEETNAME]
    SOURCE_SH = SOURCE_XL[CHOOSEN_SHEETS[0]]
    ROW_CHECK = SROW_S
    COL_CHECK = SCOL_S





POSITION = 0
for _ in CHOOSEN_SHEETS:

    print("copy of ", _)

    sheetcheck(SOURCE_XL[CHOOSEN_SHEETS[POSITION]], ROW_CHECK, COL_CHECK)
    SROW_E = sheetcheck(SOURCE_XL[CHOOSEN_SHEETS[POSITION]], ROW_CHECK, COL_CHECK)[1] - 1


    sheetcopy(SOURCE_XL[CHOOSEN_SHEETS[POSITION]], SROW_S, SROW_E, SCOL_S, SCOL_E, DROW, DCOL)
    DROW = sheetcopy(SOURCE_XL[CHOOSEN_SHEETS[POSITION]], SROW_S, SROW_E, SCOL_S, SCOL_E, DROW, DCOL)[5]
    print("next copy start from row ", DROW)

    POSITION += 1




DEST_XL.save(D_FILENAME+D_FILEEXTENSION)